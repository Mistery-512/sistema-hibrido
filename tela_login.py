import threading
import telebot
from datetime import datetime
from os import system
import customtkinter as ctk
from tkinter import *
from PIL import Image
from customtkinter import CTkImage
import os
import pathlib  # Para manipular arquivos Excel
import openpyxl
from openpyxl import load_workbook, Workbook
from tkinter import messagebox

system('cls')  # Limpa o terminal

# Caminho da pasta
pastaApp = os.path.dirname(__file__)

# Função para carregar e-mails cadastrados do Excel
def carregar_emails(arquivo_excel):
    emails = set()
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1):  # Supondo que os e-mails estão na primeira coluna
            emails.add(row[0].value)
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
    return emails

# Função para verificar o e-mail digitado
def verificar_email(event=None):
    email_digitado = username_entry.get()  # Obtém o e-mail digitado da Entry
    if not email_digitado.strip():  # Campo vazio
        mensagem_label.pack_forget()  # Esconde a mensagem de erro
        esconder_campos_senha()  # Esconde os campos de senha
    elif email_digitado in emails_cadastrados:  # E-mail cadastrado
        mensagem_label.pack_forget()  # Esconde a mensagem de erro
        mostrar_campos_senha()  # Exibe os campos de senha
    else:  # E-mail não cadastrado
        mensagem_label.pack()  # Exibe a mensagem de erro
        esconder_campos_senha()  # Esconde os campos de senha

# Funções para exibir e ocultar os campos de senha
def mostrar_campos_senha():
    senha_label.pack(pady=(10, 5))
    senha_entry.pack(pady=5)

def esconder_campos_senha():
    senha_label.pack_forget()
    senha_entry.pack_forget()

# Função para enviar a solicitação de cadastro via Telegram
def enviar_solicitacao(email):
    agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    mensagem = (
        f"Solicitação de cadastro recebida:\n\n"
        f"E-mail: {email}\n"
        f"Data/Hora: {agora}\n\n"
        f"Responda com:\n/sim - Para autorizar\n/nao - Para recusar"
    )
    try:
        bot.send_message(CHAT_ID, mensagem)
        print("Mensagem de solicitação enviada com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar mensagem de solicitação: {e}")

# Bot do Telegram
BOT_TOKEN = 'your token api'
CHAT_ID = 'your chat id'
bot = telebot.TeleBot(BOT_TOKEN)
# '7854524030'

# Função para verificar a resposta do bot
@bot.message_handler(func=lambda message: message.text.lower() in ['sim', 'não', 'nao'])
def autorizar(mensagem):
    email_digitado = username_entry.get()

    if mensagem.text.lower() == 'sim':
        # Adicionar e-mail à planilha
        caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
        if not caminho_planilha.exists():
            nova_planilha = Workbook()
            folha = nova_planilha.active
            folha['A1'] = 'E-mails'
            nova_planilha.save(caminho_planilha)

        planilha = load_workbook(caminho_planilha)
        folha = planilha.active
        folha.cell(column=1, row=folha.max_row + 1, value=email_digitado)
        planilha.save(caminho_planilha)

        # Responde ao usuário no Telegram
        bot.reply_to(mensagem, "Cadastro autorizado! E-mail foi adicionado à planilha.")
        print(f"E-mail {email_digitado} cadastrado com sucesso.")
        
        # Atualiza a interface com a confirmação do cadastro
        atualizar_interface(mensagem_label.pack_forget)
        atualizar_interface(mostrar_campos_senha)

    elif mensagem.text.lower() in ['não', 'nao']:
        bot.reply_to(mensagem, "Cadastro recusado. Nenhuma ação será tomada.")
        print(f"Cadastro do e-mail {email_digitado} recusado.")
        
        # Atualiza a interface com a recusa do cadastro
        atualizar_interface(mensagem_label.pack)
        atualizar_interface(esconder_campos_senha)

# Função para rodar o bot do Telegram em uma thread separada
def run_bot():
    bot.polling(none_stop=True)

# Função para solicitar cadastro
def solicitar_cadastro():
    email = username_entry.get()
    if email.strip():  # Verifica se o e-mail não está vazio
        enviar_solicitacao(email)
        print("Enviando solicitação...")
    else:
        messagebox.showerror("Erro", "Digite um e-mail válido para solicitar cadastro.")

bg_imagem_path = None
bg_image_pil = None
bg_image = None
bg_label = None
frame_sobreposto = None
logo_image = None
logo_label = None
username_label = None
username_entry = None
mensagem_label = None
senha_label = None
senha_entry = None
solicitar_button = None


# Função para iniciar a interface gráfica
# Função para iniciar a interface gráfica
def iniciar_interface():
    global app, bg_image, bg_image_pil, bg_imagem_path, bg_label, frame_sobreposto, logo_image, logo_label, username_label, username_entry, mensagem_label, senha_label, senha_entry, solicitar_button
    
    app = Tk()  # Propriedades da janela Tkinter
    app.title('Imagem no Python')  # Título da janela
    app.state('zoomed')  # Deixa a janela maximizada

    # Carregar a imagem de fundo com o Pillow
    bg_image_path = os.path.join(pastaApp, r'imagens/image_background_login.png')  # Caminho da imagem
    bg_image_pil = Image.open(bg_image_path)  # Carrega a imagem usando Pillow
    # Convertendo a imagem para o formato CTkImage
    bg_image = CTkImage(light_image=bg_image_pil, size=(app.winfo_screenwidth(), app.winfo_screenheight()))

    # Adicionando a imagem de fundo diretamente na janela
    bg_label = ctk.CTkLabel(
        app,
        image=bg_image,
        text="",
    )
    bg_label.place(
        relx=0.5,
        rely=0.5,
        anchor="center",
    )

    # Criar o frame sobreposto à imagem
    frame_sobreposto = ctk.CTkFrame(
        app,
        fg_color='#2d2e39',
        height=0,
        width=0,
    )
    frame_sobreposto.pack(
        side=TOP,
        padx=50,
        pady=(200, 0),
    )

    # Interface de login dentro do frame sobreposto
    logo_image = CTkImage(light_image=Image.open("backup projeto sistema\outroocoisa\imagens\logo_login_icon.png"), size=(100, 100))
    logo_label = ctk.CTkLabel(
        frame_sobreposto,
        image=logo_image,
        text="",
    )
    logo_label.pack(
        pady=(30, 5),
        padx=30,
        anchor="center",
    )

    # Campo Username
    username_label = ctk.CTkLabel(
        frame_sobreposto,
        text="User Email:",
        text_color='white',
        font=('Arial', 20, 'bold'),
    )
    username_label.pack(
        pady=(10, 5),
    )

    username_entry = ctk.CTkEntry(
        frame_sobreposto,
        placeholder_text='Insira um E-mail válido',
        width=250,
        fg_color='#6d6c89',
        justify="center",
        text_color='white',
        font=('arial', 18, 'bold'),
    )
    username_entry.pack(
        pady=5,
        padx=10,
    )
    # Bind da Entry para chamar a função verificar_email ao digitar
    username_entry.bind("<KeyRelease>", lambda event: verificar_email())

    # Label de mensagem para e-mails não cadastrados
    mensagem_label = ctk.CTkLabel(
        frame_sobreposto,
        text="E-mail de usuário não cadastrado",
        text_color="red",
        font=('Arial', 16, 'bold'),
    )
    mensagem_label.pack_forget()

    # Campos de senha (inicialmente ocultos)
    senha_label = ctk.CTkLabel(
        frame_sobreposto,
        text="Password:",
        text_color='white',
        font=('Arial', 20, 'bold'),
    )
    senha_entry = ctk.CTkEntry(
        frame_sobreposto,
        placeholder_text='Insira sua senha',
        width=250,
        fg_color='#6d6c89',
        justify="center",
        text_color='white',
        font=('arial', 18, 'bold'),
    )

    # Botão para solicitar cadastro
    solicitar_button = ctk.CTkButton(
        frame_sobreposto,
        text="Solicitar Cadastro",
        command=lambda: solicitar_cadastro(),
        width=250,
        height=40,
        fg_color="#4CAF50",  # Cor verde
        text_color="white",
        font=('Arial', 18, 'bold'),
    )
    solicitar_button.pack(pady=20)

    # Iniciar o loop da interface gráfica
    app.mainloop()


# Função para atualizar a interface de forma thread-safe
def atualizar_interface(funcao, *args):
    app.after(0, funcao, *args)

# Carregar emails cadastrados
emails_cadastrados = carregar_emails('backup projeto sistema\outroocoisa\emails_cadastrados.xlsx')

# Iniciar a interface gráfica diretamente no thread principal
iniciar_interface()

# Rodar o bot do Telegram em outra thread
thread_bot = threading.Thread(target=run_bot)
thread_bot.start()


# import threading
# import telebot
# from datetime import datetime
# from os import system
# import customtkinter as ctk
# from tkinter import *
# from PIL import Image
# from customtkinter import CTkImage
# import os
# import pathlib  # Para manipular arquivos Excel
# import openpyxl
# from openpyxl import load_workbook, Workbook
# from tkinter import messagebox

# system('cls')  # Limpa o terminal

# # Caminho da pasta
# pastaApp = os.path.dirname(__file__)

# # Função para carregar e-mails cadastrados do Excel
# def carregar_emails(arquivo_excel):
#     emails = set()
#     try:
#         wb = load_workbook(arquivo_excel)
#         ws = wb.active
#         for row in ws.iter_rows(min_row=2, max_col=1):  # Supondo que os e-mails estão na primeira coluna
#             emails.add(row[0].value)
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#     return emails

# # Função para verificar o e-mail digitado
# def verificar_email(event=None, username_entry=None, mensagem_label=None, emails_cadastrados=None):
#     email_digitado = username_entry.get()
#     if not email_digitado.strip():  # Campo vazio
#         mensagem_label.pack_forget()
#         esconder_campos_senha()
#     elif email_digitado in emails_cadastrados:
#         mensagem_label.pack_forget()
#         mostrar_campos_senha()
#     else:
#         mensagem_label.pack()
#         esconder_campos_senha()

# # Funções para exibir e ocultar os campos de senha
# def mostrar_campos_senha(senha_label, senha_entry):
#     senha_label.pack(pady=(10, 5))
#     senha_entry.pack(pady=5)

# def esconder_campos_senha(senha_label, senha_entry):
#     senha_label.pack_forget()
#     senha_entry.pack_forget()

# # Função para enviar a solicitação de cadastro via Telegram
# def enviar_solicitacao(email):
#     agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
#     mensagem = (
#         f"Solicitação de cadastro recebida:\n\n"
#         f"E-mail: {email}\n"
#         f"Data/Hora: {agora}\n\n"
#         f"Responda com:\n/sim - Para autorizar\n/nao - Para recusar"
#     )
#     try:
#         bot.send_message(CHAT_ID, mensagem)
#         print("Mensagem de solicitação enviada com sucesso!")
#     except Exception as e:
#         print(f"Erro ao enviar mensagem de solicitação: {e}")

# # Bot do Telegram
# BOT_TOKEN = 'your token api'
# CHAT_ID = 'your chat id'
# bot = telebot.TeleBot(BOT_TOKEN)

# # Função para verificar a resposta do bot
# @bot.message_handler(func=lambda message: message.text.lower() in ['sim', 'não', 'nao'])
# def autorizar(mensagem, username_entry):
#     email_digitado = username_entry.get()

#     if mensagem.text.lower() == 'sim':
#         # Adicionar e-mail à planilha
#         caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
#         if not caminho_planilha.exists():
#             nova_planilha = Workbook()
#             folha = nova_planilha.active
#             folha['A1'] = 'E-mails'
#             nova_planilha.save(caminho_planilha)

#         planilha = load_workbook(caminho_planilha)
#         folha = planilha.active
#         folha.cell(column=1, row=folha.max_row + 1, value=email_digitado)
#         planilha.save(caminho_planilha)

#         bot.reply_to(mensagem, "Cadastro autorizado! E-mail foi adicionado à planilha.")
#         print(f"E-mail {email_digitado} cadastrado com sucesso.")

#     elif mensagem.text.lower() in ['não', 'nao']:
#         bot.reply_to(mensagem, "Cadastro recusado. Nenhuma ação será tomada.")
#         print(f"Cadastro do e-mail {email_digitado} recusado.")

# # Função para rodar o bot do Telegram em uma thread separada
# def run_bot():
#     bot.polling(none_stop=True)

# # Função para solicitar cadastro
# def solicitar_cadastro(username_entry):
#     email = username_entry.get()
#     if email.strip():  # Verifica se o e-mail não está vazio
#         enviar_solicitacao(email)
#     else:
#         messagebox.showerror("Erro", "Digite um e-mail válido para solicitar cadastro.")

# class appinterface:
#     def __init__(self, master, emails_cadastrados):
#         self.master = master
#         self.emails_cadastrados = emails_cadastrados  # Passa emails cadastrados para a classe
#         self.bg_imagem_path = None
#         self.bg_image_pil = None
#         self.bg_image = None
#         self.bg_label = None
#         self.frame_sobreposto = None
#         self.logo_image = None
#         self.logo_label = None
#         self.username_label = None
#         self.username_entry = None
#         self.mensagem_label = None
#         self.senha_label = None
#         self.senha_entry = None
#         self.solicitar_button = None
    
#     def iniciar_interface(self):
#         # Carregar a imagem de fundo com o Pillow
#         self.bg_image_path = os.path.join(pastaApp, r'imagens/image_background_login.png')  # Caminho da imagem
#         self.bg_image_pil = Image.open(self.bg_image_path)  # Carrega a imagem usando Pillow
#         # Convertendo a imagem para o formato CTkImage
#         self.bg_image = CTkImage(light_image=self.bg_image_pil, size=(app.winfo_screenwidth(), app.winfo_screenheight()))

#         # Adicionando a imagem de fundo diretamente na janela
#         self.bg_label = ctk.CTkLabel(
#             app,
#             image=self.bg_image,
#             text="",
#         )
#         self.bg_label.place(
#             relx=0.5,
#             rely=0.5,
#             anchor="center",
#         )

#         # Criar o frame sobreposto à imagem
#         self.frame_sobreposto = ctk.CTkFrame(
#             app,
#             fg_color='#2d2e39',
#             height=0,
#             width=0,
#         )
#         self.frame_sobreposto.pack(
#             side=TOP,
#             padx=50,
#             pady=(200, 0),
#         )

#         # Interface de login dentro do frame sobreposto
#         self.logo_image = CTkImage(light_image=Image.open("imagens/logo_login_icon.png"), size=(100, 100))
#         self.logo_label = ctk.CTkLabel(
#             self.frame_sobreposto,
#             image=self.logo_image,
#             text="",
#         )
#         self.logo_label.pack(
#             pady=(30, 5),
#             padx=30,
#             anchor="center",
#         )

#         # Campo Username
#         self.username_label = ctk.CTkLabel(
#             self.frame_sobreposto,
#             text="User Email:",
#             text_color='white',
#             font=('Arial', 20, 'bold'),
#         )
#         self.username_label.pack(
#             pady=(10, 5),
#         )

#         self.username_entry = ctk.CTkEntry(
#             self.frame_sobreposto,
#             placeholder_text='Insira um E-mail válido',
#             width=250,
#             fg_color='#6d6c89',
#             justify="center",
#             text_color='white',
#             font=('arial', 18, 'bold'),
#         )
#         self.username_entry.pack(
#             pady=5,
#             padx=10,
#         )
#         self.username_entry.bind("<KeyRelease>", lambda event: verificar_email(event, username_entry=self.username_entry, mensagem_label=self.mensagem_label, emails_cadastrados=self.emails_cadastrados))

#         # Label de mensagem para e-mails não cadastrados
#         self.mensagem_label = ctk.CTkLabel(
#             self.frame_sobreposto,
#             text="E-mail de usuário não cadastrado",
#             text_color="red",
#             font=('Arial', 16, 'bold'),
#         )
#         self.mensagem_label.pack_forget()

#         # Campos de senha (inicialmente ocultos)
#         self.senha_label = ctk.CTkLabel(
#             self.frame_sobreposto,
#             text="Password:",
#             text_color="white",
#             font=("Arial", 18, 'bold'),
#         )
#         self.senha_entry = ctk.CTkEntry(
#             self.frame_sobreposto,
#             placeholder_text="********",
#             width=250,
#             fg_color='#6d6c89',
#             justify="center",
#             text_color='white',
#             font=('Arial', 18, 'bold'),
#             show="*",
#         )

#         # Botão para solicitar cadastro via Telegram
#         self.solicitar_button = ctk.CTkButton(
#             self.frame_sobreposto,
#             text="Solicitar Cadastro",
#             width=200,
#             height=40,
#             command=solicitar_cadastro,
#         )
#         self.solicitar_button.pack(
#             pady=20,
#         )

# # Carregar e-mails cadastrados
# emails_cadastrados = carregar_emails('emails_cadastrados.xlsx')

# # Criar e iniciar a interface
# app = Tk()
# app.title("Sistema de login")
# app.state("zoomed")
# interface = appinterface(app, emails_cadastrados)
# interface.iniciar_interface()

'''from os import system
import customtkinter as ctk
from tkinter import *
from PIL import Image
from customtkinter import CTkImage
import os
import openpyxl
import pathlib  # Para manipular arquivos Excel
from openpyxl import load_workbook, Workbook
from tkinter import messagebox

system('cls')  # Limpa o terminal

# Caminho da pasta
pastaApp = os.path.dirname(__file__)


# Função para carregar e-mails cadastrados do Excel
def carregar_emails(arquivo_excel):
    emails = set()
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1):  # Supondo que os e-mails estão na primeira coluna
            emails.add(row[0].value)
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
    return emails


# Função para verificar o e-mail digitado
def verificar_email(event=None):
    email_digitado = username_entry.get()
    if not email_digitado.strip():  # Campo vazio
        mensagem_label.pack_forget()
        esconder_campos_senha()
    elif email_digitado in emails_cadastrados:
        mensagem_label.pack_forget()
        mostrar_campos_senha()
    else:
        mensagem_label.pack()
        esconder_campos_senha()

# Funções para exibir e ocultar os campos de senha
def mostrar_campos_senha():
    senha_label.pack(pady=(10, 5))
    senha_entry.pack(pady=5)

def esconder_campos_senha():
    senha_label.pack_forget()
    senha_entry.pack_forget()

# Configuração da janela principal
app = Tk()  # Propriedades da janela Tkinter
app.title('Imagem no Python')  # Título da janela
app.state('zoomed')  # Deixa a janela maximizada

# Carregar a imagem de fundo com o Pillow
bg_image_path = os.path.join(pastaApp, r'imagens/image_background_login.png')  # Caminho da imagem
bg_image_pil = Image.open(bg_image_path)  # Carrega a imagem usando Pillow
# Convertendo a imagem para o formato CTkImage
bg_image = CTkImage(light_image=bg_image_pil, size=(app.winfo_screenwidth(), app.winfo_screenheight()))

# Adicionando a imagem de fundo diretamente na janela
bg_label = ctk.CTkLabel(
    app,
    image=bg_image,
    text="",
)
bg_label.place(
    relx=0.5,
    rely=0.5,
    anchor="center",
)

# Criar o frame sobreposto à imagem
frame_sobreposto = ctk.CTkFrame(
    app,
    fg_color='#2d2e39',
    height=0,
    width=0,
)
frame_sobreposto.pack(
    side=TOP,
    padx=50,
    pady=(200, 0),
)

# Interface de login dentro do frame sobreposto
logo_image = CTkImage(light_image=Image.open("imagens/logo_login_icon.png"), size=(100, 100))
logo_label = ctk.CTkLabel(
    frame_sobreposto,
    image=logo_image,
    text="",
)
logo_label.pack(
    pady=(30, 5),
    padx=30,
    anchor="center",
)

# Campo Username
username_label = ctk.CTkLabel(
    frame_sobreposto,
    text="User Email:",
    text_color='white',
    font=('Arial', 20, 'bold'),
)
username_label.pack(
    pady=(10, 5),
)

username_entry = ctk.CTkEntry(
    frame_sobreposto,
    placeholder_text='Insira um E-mail válido',
    width=250,
    fg_color='#6d6c89',
    justify="center",
    text_color='white',
    font=('arial', 18, 'bold'),
)
username_entry.pack(
    pady=5,
    padx=10,
)
username_entry.bind("<KeyRelease>", verificar_email)

# Label de mensagem para e-mails não cadastrados
mensagem_label = ctk.CTkLabel(
    frame_sobreposto,
    text="E-mail de usuário não cadastrado",
    text_color="red",
    font=('Arial', 16, 'bold'),
)
mensagem_label.pack_forget()

# Campos de senha (inicialmente ocultos)
senha_label = ctk.CTkLabel(
    frame_sobreposto,
    text="Password:",
    text_color='white',
    font=('Arial', 20, 'bold'),
)
senha_entry = ctk.CTkEntry(
    frame_sobreposto,
    placeholder_text='Insira sua senha',
    width=250,
    fg_color='#6d6c89',
    justify="center",
    text_color='white',
    font=('arial', 18, 'bold'),
)

# Carregar emails cadastrados
emails_cadastrados = carregar_emails("emails_cadastrados.xlsx")

# Iniciar a interface gráfica
app.mainloop()'''


# import customtkinter as ctk
# from tkinter import *
# from PIL import Image
# from customtkinter import CTkImage
# import os
# import openpyxl
# import pathlib  # Para manipular arquivos Excel
# from openpyxl import load_workbook, Workbook
# from tkinter import messagebox
# from telegram import Bot
# from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext
# import asyncio
# from os import system

# system('cls')  # Limpa o terminal

# # Caminho da pasta
# pastaApp = os.path.dirname(__file__)

# # Configuração Telegram
# TELEGRAM_TOKEN = "your token api"
# CHAT_ID = "your chat id"  # Para obter o chat_id, envie uma mensagem para o bot e acesse https://api.telegram.org/bot<seu_token>/getUpdates

# bot = Bot(token=TELEGRAM_TOKEN)

# # Função para carregar e-mails cadastrados do Excel
# def carregar_emails(arquivo_excel):
#     emails = set()
#     try:
#         wb = load_workbook(arquivo_excel)
#         ws = wb.active
#         for row in ws.iter_rows(min_row=2, max_col=1):  # Supondo que os e-mails estão na primeira coluna
#             emails.add(row[0].value)
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#     return emails

# # # Função para enviar a solicitação via Telegram
# # async def solicitar_cadastro_telegram():
# #     email_digitado = username_entry.get()
# #     print(f"Email digitado: {email_digitado}")  # Verifique se o e-mail está sendo capturado corretamente
# #     if not email_digitado.strip():
# #         mensagem_label.configure(text="Insira um e-mail válido!", text_color="red")
# #         mensagem_label.pack()
# #         return

# # Função para enviar a solicitação via Telegram
# async def solicitar_cadastro_telegram():
#     email_digitado = username_entry.get()
#     if not email_digitado.strip():
#         mensagem_label.configure(text="Insira um e-mail válido!", text_color="red")
#         mensagem_label.pack()
#         return
#     try:
#         # Envia uma mensagem de teste para o seu chat no Telegram de forma assíncrona
#         await bot.send_message(chat_id=CHAT_ID, text="Testando bot!")
#         print("Mensagem enviada com sucesso!")
#     except Exception as e:
#         print(f"Erro ao enviar mensagem: {e}")

#     # Enviar solicitação por notificação Telegram
#     try:
#         mensagem = f"E-mail: {email_digitado} Responda SIM para aprovar ou NÃO para recusar."
#         await bot.send_message(chat_id=CHAT_ID, text=mensagem)  # Agora usando await
#         mensagem_label.configure(
#             text="Solicitação enviada para aprovação!",
#             text_color="green"
#         )
#         mensagem_label.pack()
#         # Retorna ao comportamento original após 3 segundos
#         mensagem_label.after(3000, verificar_email)
#     except Exception as e:
#         mensagem_label.configure(
#             text=f"Erro ao enviar solicitação: {e}",
#             text_color="red"
#         )
#         mensagem_label.pack()
#         # Retorna ao comportamento original após 3 segundos
#         mensagem_label.after(3000, verificar_email)

# # Função para verificar respostas do Telegram
# async def verificar_respostas_telegram(update, context):
#     resposta = update.message.text.lower().strip()
#     email_digitado = username_entry.get()  # Agora o código pode acessar username_entry

#     if "sim" in resposta:
#         # Adicionar e-mail à planilha
#         caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
#         if not caminho_planilha.exists():
#             nova_planilha = Workbook()
#             folha = nova_planilha.active
#             folha['A1'] = 'E-mails'
#             nova_planilha.save(caminho_planilha)

#         planilha = load_workbook(caminho_planilha)
#         folha = planilha.active
#         folha.cell(column=1, row=folha.max_row + 1, value=email_digitado)
#         planilha.save(caminho_planilha)

#         messagebox.showinfo('Sistema', 'E-mail aprovado e cadastrado com sucesso.')
#     elif "não" in resposta or "nao" in resposta:
#         messagebox.showerror('Sistema', 'E-mail Recusado. Nenhuma ação será tomada.')
#     else:
#         print("Resposta inválida recebida.")

# # Função para iniciar o bot e lidar com comandos
# async def start_bot():
#     application = Application.builder().token(TELEGRAM_TOKEN).build()
    
#     # Adiciona handlers para as mensagens recebidas
#     application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, verificar_respostas_telegram))

#     # Inicia o bot
#     await application.run_polling()

# # Função para verificar o e-mail digitado
# def verificar_email(event=None):
#     email_digitado = username_entry.get()
#     if not email_digitado.strip():  # Campo vazio
#         mensagem_label.pack_forget()
#         esconder_campos_senha()
#     elif email_digitado in emails_cadastrados:
#         mensagem_label.pack_forget()
#         mostrar_campos_senha()
#     else:
#         mensagem_label.pack()
#         esconder_campos_senha()

# # Funções para exibir e ocultar os campos de senha
# def mostrar_campos_senha():
#     senha_label.pack(pady=(10, 5))
#     senha_entry.pack(pady=5)

# def esconder_campos_senha():
#     senha_label.pack_forget()
#     senha_entry.pack_forget()

# # Configuração da janela principal
# app = Tk()  # Propriedades da janela Tkinter
# app.title('Imagem no Python')  # Título da janela
# app.state('zoomed')  # Deixa a janela maximizada

# # Carregar a imagem de fundo com o Pillow
# bg_image_path = os.path.join(pastaApp, r'imagens/image_background_login.png')  # Caminho da imagem
# bg_image_pil = Image.open(bg_image_path)  # Carrega a imagem usando Pillow
# # Convertendo a imagem para o formato CTkImage
# bg_image = CTkImage(light_image=bg_image_pil, size=(app.winfo_screenwidth(), app.winfo_screenheight()))

# # Adicionando a imagem de fundo diretamente na janela
# bg_label = ctk.CTkLabel(
#     app,
#     image=bg_image,
#     text="",
# )
# bg_label.place(
#     relx=0.5,
#     rely=0.5,
#     anchor="center",
# )

# # Criar o frame sobreposto à imagem
# frame_sobreposto = ctk.CTkFrame(
#     app,
#     fg_color='#2d2e39',
#     height=0,
#     width=0,
# )
# frame_sobreposto.pack(
#     side=TOP,
#     padx=50,
#     pady=(200, 0),
# )

# # Interface de login dentro do frame sobreposto
# logo_image = CTkImage(light_image=Image.open("imagens/logo_login_icon.png"), size=(100, 100))
# logo_label = ctk.CTkLabel(
#     frame_sobreposto,
#     image=logo_image,
#     text="",
# )
# logo_label.pack(
#     pady=(30, 5),
#     padx=30,
#     anchor="center",
# )

# # Campo Username
# username_label = ctk.CTkLabel(
#     frame_sobreposto,
#     text="User Email:",
#     text_color='white',
#     font=('Arial', 20, 'bold'),
# )
# username_label.pack(
#     pady=(10, 5),
# )

# username_entry = ctk.CTkEntry(
#     frame_sobreposto,
#     placeholder_text='Insira um E-mail válido',
#     width=250,
#     fg_color='#6d6c89',
#     justify="center",
#     text_color='white',
#     font=('arial', 18, 'bold'),
# )
# username_entry.pack(
#     pady=5,
#     padx=10,
# )
# username_entry.bind("<KeyRelease>", verificar_email)

# botao_solicitar = ctk.CTkButton(
#     frame_sobreposto,
#     text="Solicitar Cadastro",
#     command=solicitar_cadastro_telegram,
#     fg_color="#4caf50",
#     text_color="white",
#     font=('Arial', 18, 'bold'),
# )
# botao_solicitar.pack(pady=(10, 20))

# # Label de mensagem para e-mails não cadastrados
# mensagem_label = ctk.CTkLabel(
#     frame_sobreposto,
#     text="E-mail de usuário não cadastrado",
#     text_color="red",
#     font=('Arial', 16, 'bold'),
# )
# mensagem_label.pack_forget()

# # Campos de senha (inicialmente ocultos)
# senha_label = ctk.CTkLabel(
#     frame_sobreposto,
#     text="Password:",
#     text_color='white',
#     font=('Arial', 20, 'bold'),
# )
# senha_entry = ctk.CTkEntry(
#     frame_sobreposto,
#     placeholder_text='Insira sua senha',
#     width=250,
#     fg_color='#6d6c89',
#     justify="center",
#     text_color='white',
#     font=('arial', 18, 'bold'),
# )

# # Carregar emails cadastrados
# emails_cadastrados = carregar_emails("emails_cadastrados.xlsx")

# # Iniciar o bot com o loop de eventos
# loop = asyncio.get_event_loop()
# loop.create_task(start_bot())  # Corrigido para não gerar mais avisos

# # Iniciar a interface gráfica
# app.mainloop()




'''from os import system
import customtkinter as ctk
from tkinter import *
from PIL import Image
from customtkinter import CTkImage
import os
import openpyxl
import pathlib  # Para manipular arquivos Excel
from openpyxl import load_workbook, Workbook
from tkinter import messagebox
from telegram import Bot
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext
import threading



system('cls')  # Limpa o terminal

# Caminho da pasta
pastaApp = os.path.dirname(__file__)

# Configuração Telegram
TELEGRAM_TOKEN = "your token api"
CHAT_ID = "your chat id"  # Para obter o chat_id, envie uma mensagem para o bot e acesse https://api.telegram.org/bot<seu_token>/getUpdates

bot = Bot(token=TELEGRAM_TOKEN)

# Função para carregar e-mails cadastrados do Excel
def carregar_emails(arquivo_excel):
    emails = set()
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1):  # Supondo que os e-mails estão na primeira coluna
            emails.add(row[0].value)
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
    return emails

# Função para enviar a solicitação via Telegram
def solicitar_cadastro_telegram():
    email_digitado = username_entry.get()
    print(f"Email digitado: {email_digitado}")  # Verifique se o e-mail está sendo capturado corretamente
    if not email_digitado.strip():
        mensagem_label.configure(text="Insira um e-mail válido!", text_color="red")
        mensagem_label.pack()
        return

    # Enviar solicitação por notificação Telegram
    try:
        mensagem = f"E-mail: {email_digitado} Responda SIM para aprovar ou NÃO para recusar."
        bot.send_message(chat_id=CHAT_ID, text=mensagem)  # Envia a mensagem para o chat
        mensagem_label.configure(
            text="Solicitação enviada para aprovação!",
            text_color="green"
        )
        mensagem_label.pack()
        # Retorna ao comportamento original após 3 segundos
        mensagem_label.after(3000, verificar_email)
    except Exception as e:
        mensagem_label.configure(
            text=f"Erro ao enviar solicitação: {e}",
            text_color="red"
        )
        mensagem_label.pack()
        # Retorna ao comportamento original após 3 segundos
        mensagem_label.after(3000, verificar_email)


# Função para verificar respostas do Telegram
def verificar_respostas_telegram(update, context, username_entry):
    resposta = update.message.text.lower().strip()
    email_digitado = username_entry.get()  # Agora o código pode acessar username_entry

    if "sim" in resposta:
        # Adicionar e-mail à planilha
        caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
        if not caminho_planilha.exists():
            nova_planilha = Workbook()
            folha = nova_planilha.active
            folha['A1'] = 'E-mails'
            nova_planilha.save(caminho_planilha)

        planilha = load_workbook(caminho_planilha)
        folha = planilha.active
        folha.cell(column=1, row=folha.max_row + 1, value=email_digitado)
        planilha.save(caminho_planilha)

        messagebox.showinfo('Sistema', 'E-mail aprovado e cadastrado com sucesso.')
    elif "não" in resposta or "nao" in resposta:
        messagebox.showerror('Sistema', 'E-mail Recusado. Nenhuma ação será tomada.')
    else:
        print("Resposta inválida recebida.")


# Função para iniciar o bot e lidar com comandos
def start_bot():
    application = Application.builder().token(TELEGRAM_TOKEN).build()
    
    # Adiciona handlers para as mensagens recebidas
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, lambda update, context: verificar_respostas_telegram(update, context, username_entry)))

    # Inicia o bot
    application.run_polling()



# Função para verificar o e-mail digitado
def verificar_email(event=None):
    email_digitado = username_entry.get()
    if not email_digitado.strip():  # Campo vazio
        mensagem_label.pack_forget()
        esconder_campos_senha()
    elif email_digitado in emails_cadastrados:
        mensagem_label.pack_forget()
        mostrar_campos_senha()
    else:
        mensagem_label.pack()
        esconder_campos_senha()

# Funções para exibir e ocultar os campos de senha
def mostrar_campos_senha():
    senha_label.pack(pady=(10, 5))
    senha_entry.pack(pady=5)

def esconder_campos_senha():
    senha_label.pack_forget()
    senha_entry.pack_forget()

# Configuração da janela principal
app = Tk()  # Propriedades da janela Tkinter
app.title('Imagem no Python')  # Título da janela
app.state('zoomed')  # Deixa a janela maximizada

# Função para rodar o bot em um thread separado
def run_telegram_bot():
    start_bot()  # Inicia o bot Telegram

# Rodar o bot em um thread separado
bot_thread = threading.Thread(target=run_telegram_bot)
bot_thread.daemon = True  # Garante que o thread será encerrado quando o programa principal for fechado
bot_thread.start()

# Carregar a imagem de fundo com o Pillow
bg_image_path = os.path.join(pastaApp, r'imagens/image_background_login.png')  # Caminho da imagem
bg_image_pil = Image.open(bg_image_path)  # Carrega a imagem usando Pillow
# Convertendo a imagem para o formato CTkImage
bg_image = CTkImage(light_image=bg_image_pil, size=(app.winfo_screenwidth(), app.winfo_screenheight()))

# Adicionando a imagem de fundo diretamente na janela
bg_label = ctk.CTkLabel(
    app,
    image=bg_image,
    text="",
)
bg_label.place(
    relx=0.5,
    rely=0.5,
    anchor="center",
)

# Criar o frame sobreposto à imagem
frame_sobreposto = ctk.CTkFrame(
    app,
    fg_color='#2d2e39',
    height=0,
    width=0,
)
frame_sobreposto.pack(
    side=TOP,
    padx=50,
    pady=(200, 0),
)

# Interface de login dentro do frame sobreposto
logo_image = CTkImage(light_image=Image.open("imagens/logo_login_icon.png"), size=(100, 100))
logo_label = ctk.CTkLabel(
    frame_sobreposto,
    image=logo_image,
    text="",
)
logo_label.pack(
    pady=(30, 5),
    padx=30,
    anchor="center",
)

# Campo Username
username_label = ctk.CTkLabel(
    frame_sobreposto,
    text="User Email:",
    text_color='white',
    font=('Arial', 20, 'bold'),
)
username_label.pack(
    pady=(10, 5),
)

username_entry = ctk.CTkEntry(
    frame_sobreposto,
    placeholder_text='Insira um E-mail válido',
    width=250,
    fg_color='#6d6c89',
    justify="center",
    text_color='white',
    font=('arial', 18, 'bold'),
)
username_entry.pack(
    pady=5,
    padx=10,
)
username_entry.bind("<KeyRelease>", verificar_email)

botao_solicitar = ctk.CTkButton(
    frame_sobreposto,
    text="Solicitar Cadastro",
    command=solicitar_cadastro_telegram,
    fg_color="#4caf50",
    text_color="white",
    font=('Arial', 18, 'bold'),
)
botao_solicitar.pack(pady=(10, 20))

# Label de mensagem para e-mails não cadastrados
mensagem_label = ctk.CTkLabel(
    frame_sobreposto,
    text="E-mail de usuário não cadastrado",
    text_color="red",
    font=('Arial', 16, 'bold'),
)
mensagem_label.pack_forget()

# Campos de senha (inicialmente ocultos)
senha_label = ctk.CTkLabel(
    frame_sobreposto,
    text="Password:",
    text_color='white',
    font=('Arial', 20, 'bold'),
)
senha_entry = ctk.CTkEntry(
    frame_sobreposto,
    placeholder_text='Digite sua senha',
    width=250,
    fg_color='#6d6c89',
    justify="center",
    text_color='white',
    font=('arial', 18, 'bold'),
    show=("•"),
)


# Carregar e-mails cadastrados
emails_cadastrados = carregar_emails("emails_cadastrados.xlsx")  # Substitua pelo caminho correto do arquivo

# Loop da interface
app.mainloop()'''





""" 
from tkinter import *
from os import system
import customtkinter as ctk
from PIL import Image
from customtkinter import CTkImage
import os
import openpyxl
import pathlib  # Para manipular arquivos Excel
from openpyxl import load_workbook, Workbook
from tkinter import messagebox
from pushbullet import Pushbullet  # Biblioteca para Pushbullet

system('cls')  # Limpa o terminal

# Caminho da pasta
pastaApp = os.path.dirname(__file__)

# Configuração Pushbullet
PUSHBULLET_API_KEY = "Your Token API"  # Insira sua chave da API do Pushbullet
pushbullet = Pushbullet(PUSHBULLET_API_KEY)

# Função para carregar e-mails cadastrados do Excel
def carregar_emails(arquivo_excel):
    emails = set()
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1):  # Supondo que os e-mails estão na primeira coluna
            emails.add(row[0].value)
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
    return emails

# Função para solicitar cadastro via SMS
def solicitar_cadastro():
    email_digitado = username_entry.get()
    if not email_digitado.strip():
        mensagem_label.configure(text="Insira um e-mail válido!", text_color="red")
        mensagem_label.pack()
        return

    # Enviar solicitação por SMS via Pushbullet
    try:
        mensagem = f"Solicitação de cadastro do e-mail: {email_digitado}. Responda SIM para aprovar ou NÃO para recusar."
        pushbullet.push_sms(pushbullet.devices[0], "+55 81 985252203", mensagem)  # Altere para o número correto

        mensagem_label.configure(
            text="Solicitação enviada para aprovação!",
            text_color="green"
        )
        mensagem_label.pack()
    except Exception as e:
        mensagem_label.configure(
            text=f"Erro ao enviar solicitação: {e}",
            text_color="red"
        )
        mensagem_label.pack()

# Função para verificar respostas de SMS
def verificar_respostas_sms():
    try:
        mensagens = pushbullet.get_pushes(limit=5)  # Obtém as últimas 5 mensagens
        for mensagem in mensagens:
            if "Solicitação de cadastro" in mensagem.get("body", ""):
                processar_resposta_sms(mensagem)
    except Exception as e:
        print(f"Erro ao verificar respostas: {e}")

# Função para processar respostas de SMS
def processar_resposta_sms(mensagem):
    corpo = mensagem["body"].lower().strip()
    email_digitado = username_entry.get()

    if "sim" in corpo:
        # Adicionar e-mail à planilha
        caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
        if not caminho_planilha.exists():
            nova_planilha = Workbook()
            folha = nova_planilha.active
            folha['A1'] = 'E-mails'
            nova_planilha.save(caminho_planilha)

        planilha = load_workbook(caminho_planilha)
        folha = planilha.active
        folha.cell(column=1, row=folha.max_row + 1, value=email_digitado)
        planilha.save(caminho_planilha)

        messagebox.showinfo('Sistema', 'E-mail aprovado e cadastrado com sucesso.')
    elif "não" in corpo or "nao" in corpo:
        messagebox.showerror('Sistema', 'E-mail Recusado. Nenhuma ação será tomada.')
    else:
        print("Resposta inválida recebida.")

# Função para verificar o e-mail digitado
def verificar_email(event=None):
    email_digitado = username_entry.get()
    if not email_digitado.strip():  # Campo vazio
        mensagem_label.pack_forget()
        esconder_campos_senha()
    elif email_digitado in emails_cadastrados:
        mensagem_label.pack_forget()
        mostrar_campos_senha()
    else:
        mensagem_label.pack()
        esconder_campos_senha()

# Funções para exibir e ocultar os campos de senha
def mostrar_campos_senha():
    senha_label.pack(pady=(10, 5))
    senha_entry.pack(pady=5)

def esconder_campos_senha():
    senha_label.pack_forget()
    senha_entry.pack_forget()

# Configuração da interface gráfica
app = Tk()
app.title('Imagem no Python')
app.state('zoomed')

# Carregar a imagem de fundo
bg_image_path = os.path.join(pastaApp, r'imagens/image_background_login.png')
bg_image_pil = Image.open(bg_image_path)
bg_image = CTkImage(light_image=bg_image_pil, size=(app.winfo_screenwidth(), app.winfo_screenheight()))

bg_label = ctk.CTkLabel(app, image=bg_image, text="")
bg_label.place(relx=0.5, rely=0.5, anchor="center")

frame_sobreposto = ctk.CTkFrame(app, fg_color='#2d2e39', height=0, width=0)
frame_sobreposto.pack(side=TOP, padx=50, pady=(200, 0))

# Logo
logo_image = CTkImage(light_image=Image.open("imagens/logo_login_icon.png"), size=(100, 100))
logo_label = ctk.CTkLabel(frame_sobreposto, image=logo_image, text="")
logo_label.pack(pady=(30, 5), padx=30, anchor="center")

# Campo de entrada de e-mail
username_label = ctk.CTkLabel(frame_sobreposto, text="User Email:", text_color='white', font=('Arial', 20, 'bold'))
username_label.pack(pady=(10, 5))

username_entry = ctk.CTkEntry(
    frame_sobreposto,
    placeholder_text='Insira um E-mail válido',
    width=250,
    fg_color='#6d6c89',
    justify="center",
    text_color='white',
    font=('arial', 18, 'bold'),
)
username_entry.pack(pady=5, padx=10)

botao_solicitar = ctk.CTkButton(
    frame_sobreposto,
    text="Solicitar Cadastro",
    command=solicitar_cadastro,
    fg_color="#4caf50",
    text_color="white",
    font=('Arial', 18, 'bold'),
)
botao_solicitar.pack(pady=(10, 20))

mensagem_label = ctk.CTkLabel(
    frame_sobreposto, 
    text="E-mail de usuário não cadastrado",
    text_color="red",
    font=('Arial', 16, 'bold'),
)
mensagem_label.pack_forget()

# Campos de senha (inicialmente ocultos)
senha_label = ctk.CTkLabel(
    frame_sobreposto,
    text="Password:",
    text_color='white',
    font=('Arial', 20, 'bold'),
)
senha_entry = ctk.CTkEntry(
    frame_sobreposto,
    placeholder_text='Digite sua senha',
    width=250,
    fg_color='#6d6c89',
    justify="center",
    text_color='white',
    font=('arial', 18, 'bold'),
    show=("•"),
)


# Verificar SMS periodicamente
def verificar_respostas_periodicamente():
    verificar_respostas_sms()
    app.after(60000, verificar_respostas_periodicamente)

verificar_respostas_periodicamente()

# Carregar e-mails cadastrados
emails_cadastrados = carregar_emails("emails_cadastrados.xlsx")  # Substitua pelo caminho correto do arquivo

# Loop principal
app.mainloop() """







# from os import system
# import customtkinter as ctk
# from tkinter import *
# import os
# from PIL import Image
# from customtkinter import CTkImage


# system('cls')  # Limpa o terminal


# # Caminho da pasta
# pastaApp = os.path.dirname(__file__)


# # Configuração da janela principal
# app = Tk() # Propriedades da janela Tkinter
# app.title('Imagem no Python') # Título da janela
# app.state('zoomed')  # Deixa a janela maximizada


# # Carregar a imagem de fundo com o Pillow
# bg_image_path = os.path.join(pastaApp, r"imagens\image_background_login.png") # Caminho da imagem
# bg_image_pil = Image.open(bg_image_path)  # Carrega a imagem usando Pillow
# # Convertendo a imagem para o formato CTkImage
# bg_image = CTkImage(light_image=bg_image_pil,size=(app.winfo_screenwidth(), app.winfo_screenheight()))


# # Adicionando a imagem de fundo diretamente na janela
# bg_label = ctk.CTkLabel(
#                         app, 
#                         image=bg_image, 
#                         text="",
#                         )
# bg_label.place(
#                 relx=0.5, 
#                 rely=0.5, 
#                 anchor="center",
#                 )


# # Criar o frame sobreposto à imagem
# frame_sobreposto = ctk.CTkFrame(
#                                 app,
#                                 fg_color='#2d2e39',
#                                 height=0,
#                                 width=0,
#                                 )
# frame_sobreposto.pack(
#                 side=TOP, 
#                 padx=50, 
#                 pady=(200,0),
#                 )


# # Interface de login dentro do frame sobreposto


# # Adicionando a logo com fundo transparente
# logo_image = CTkImage(light_image=Image.open("imagens/logo_login_icon.png"), size=(100, 100))
# logo_label = ctk.CTkLabel(
#                         frame_sobreposto,
#                         image=logo_image, 
#                         text="",
#                         )
# logo_label.pack(
#                 pady=(30,5), 
#                 padx=30,
#                 anchor="center",
#                 )


# # Campo Username
# username_label = ctk.CTkLabel(
#                         frame_sobreposto, 
#                         text="User Email:", 
#                         text_color='white',
#                         font=('Arial', 20, 'bold'),
#                         )
# username_label.pack(
#                 pady=(10,5),
#                 )


# username_entry = ctk.CTkEntry(
#                         frame_sobreposto, 
#                         placeholder_text='Insira um E-mail válido',
#                         width=250,
#                         fg_color='#6d6c89',
#                         justify="center",
#                         text_color='white',
#                         font=('arial',18,'bold'),
#                         )
# username_entry.pack(
#                 pady=5,
#                 padx=10,
#                 )


# # Loop da interface
# app.mainloop()