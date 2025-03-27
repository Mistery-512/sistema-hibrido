import customtkinter as ctk
import telebot
import threading
import queue
from openpyxl import Workbook, load_workbook
import pathlib

# Configuração do bot do Telegram
API_KEY = "your API key"
bot = telebot.TeleBot(API_KEY)

# Fila para comunicação entre o bot e a interface
message_queue = queue.Queue()

def adicionar_email_planilha(email):
    caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
    
    # Cria a planilha se não existir
    if not caminho_planilha.exists():
        nova_planilha = Workbook()
        folha = nova_planilha.active
        folha['A1'] = 'E-mails'
        nova_planilha.save(caminho_planilha)
    
    # Abre a planilha existente
    planilha = load_workbook(caminho_planilha)
    folha = planilha.active

    # Encontra a próxima linha realmente vazia
    next_row = 1  # Começa da linha 1
    for row in folha.iter_rows(min_row=1, max_col=1, values_only=True):
        if not row[0]:  # Verifica se a célula está vazia
            break
        next_row += 1

    # Insere o e-mail na linha vazia encontrada
    folha.cell(column=1, row=next_row, value=email)
    planilha.save(caminho_planilha)

# Função do bot rodando em uma thread
def run_telegram_bot():
    @bot.message_handler(commands=['start'])
    def send_welcome(message):
        bot.reply_to(message, "Olá! Envie um e-mail para cadastrar na planilha.")
    
    @bot.message_handler(func=lambda message: True)
    def handle_message(message):
        email = message.text.strip()
        # Simples validação de e-mail
        if "@" in email and "." in email:
            adicionar_email_planilha(email)
            bot.reply_to(message, "E-mail cadastrado com sucesso!")
            # Envia a mensagem para a fila para atualizar a interface
            message_queue.put(f"E-mail cadastrado: {email}")
        else:
            bot.reply_to(message, "Formato de e-mail inválido. Tente novamente.")
    
    bot.polling()

# Classe da interface gráfica
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Bot Telegram + Interface Gráfica")
        self.geometry("600x400")
        
        self.label = ctk.CTkLabel(self, text="Aguardando mensagens...", anchor="w", width=400)
        self.label.pack(pady=10, padx=10, fill="x")
        
        self.textbox = ctk.CTkTextbox(self, width=500, height=300)
        self.textbox.pack(pady=10, padx=10)
        
        self.refresh_button = ctk.CTkButton(self, text="Atualizar Dados", command=self.update_data)
        self.refresh_button.pack(pady=10)
        
        self.update_interface()  # Inicia a atualização periódica da interface
    
    def update_data(self):
        try:
            # Lê os dados da planilha para exibição
            caminho_planilha = pathlib.Path('emails_cadastrados.xlsx')
            if caminho_planilha.exists():
                planilha = load_workbook(caminho_planilha)
                folha = planilha.active
                data = [[cell.value for cell in row] for row in folha.iter_rows()]
                display_text = "\n".join(["\t".join(map(str, row)) for row in data])
                self.textbox.delete("1.0", "end")  # Limpa o textbox
                self.textbox.insert("1.0", display_text)  # Adiciona os dados
            else:
                self.textbox.delete("1.0", "end")
                self.textbox.insert("1.0", "Nenhuma planilha encontrada.")
        except Exception as e:
            self.textbox.delete("1.0", "end")
            self.textbox.insert("1.0", f"Erro ao acessar a planilha: {e}")

    def update_interface(self):
        try:
            # Verifica se há mensagens na fila
            while not message_queue.empty():
                msg = message_queue.get()
                self.label.configure(text=msg)
        except Exception as e:
            print(f"Erro ao atualizar interface: {e}")
        
        # Agenda nova verificação
        self.after(100, self.update_interface)

# Inicia o bot em uma thread separada
bot_thread = threading.Thread(target=run_telegram_bot, daemon=True)
bot_thread.start()

# Inicia a interface gráfica
app = App()
app.mainloop()


""" import customtkinter as ctk
import telebot
import threading
import queue

# Configuração do bot do Telegram
API_KEY = "your API key"
bot = telebot.TeleBot(API_KEY)

# Fila para comunicação entre o bot e a interface
message_queue = queue.Queue()

# Função do bot rodando em uma thread
def run_telegram_bot():
    @bot.message_handler(commands=['start'])
    def send_welcome(message):
        bot.reply_to(message, "Olá! Envie uma mensagem para testar.")
    
    @bot.message_handler(func=lambda message: True)
    def echo_message(message):
        # Envia a mensagem para a fila
        message_queue.put(f"Recebido do Telegram: {message.text}")
        bot.reply_to(message, "Mensagem recebida!")
    
    bot.polling()

# Classe da interface gráfica
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Bot Telegram + Interface Gráfica")
        self.geometry("500x300")
        
        self.label = ctk.CTkLabel(self, text="Aguardando mensagens...", width=400)
        self.label.pack(pady=20)
        
        self.update_interface()
    
    def update_interface(self):
        try:
            # Verifica se há mensagens na fila
            while not message_queue.empty():
                msg = message_queue.get()
                self.label.configure(text=msg)
        except Exception as e:
            print(f"Erro ao atualizar interface: {e}")
        
        # Agenda nova verificação
        self.after(100, self.update_interface)

# Inicia o bot em uma thread separada
bot_thread = threading.Thread(target=run_telegram_bot, daemon=True)
bot_thread.start()

# Inicia a interface gráfica
app = App()
app.mainloop() """