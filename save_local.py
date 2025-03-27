import customtkinter as ctk 
from tkinter import *
from tkinter import messagebox 
import openpyxl, xlrd 
import pathlib #responsável pelos ficheiros excel
from openpyxl import Workbook 


#definindo a aparência padrão do nosso sistema de cadastro

ctk.set_appearance_mode('System')
ctk.set_default_color_theme('blue')



class App(ctk.CTk): #iniciando a janela
    def __init__(self): #variável global
        super().__init__()  #transformamos a classe init é de nível mais avançado
        self.layout_tela() #chamando a função da tela. Assim, ela irá iniciar.
        self.temas() #chamando a função de temas, para que o usuário possa trocar o tema do sistema pelo botão menu de opções.
        self.wigets()    
    
    
    def layout_tela(self): #função da tela
        self.title('Janela principal') #título da janela
        self.geometry('700x510') #tamanho da janela



    def temas(self): #Definindo o tema do sistema. O text color é uma lista para caso o usuário queira trocar de tema,
        #lb_apm = label appearence mode
        #fazendo a cor da fonte ser alterada no processo! #000 é a cor preta. #fff é o branco.
        self.lb_apm = ctk.CTkLabel(self, text = 'Tema:', bg_color = 'transparent', text_color = ['#000', '#fff']).place(x = 50, y = 430) #direcionando a label
        self.opt_apm = ctk.CTkOptionMenu(self, values = ['Light', 'Dark', 'System'], command = self.change_apm).place(x = 50, y = 460) #esses valores servem para o usuário escolher a cor do tema do sistema. Está em lista justamente para ser de múltipla escolha.    menu de opções de tema(appearence mode)
    
    
    
    
    def wigets(self): #criando os widgets do sistema. Escrevi errado para não confundir.
        frame = ctk.CTkFrame(self, width = 700, height = 50, corner_radius = 0, bg_color = 'teal', fg_color = 'teal').place(x = 0, y = 10)
        
        title = ctk.CTkLabel(frame, text = 'Sistema de Locadora Road Réquiem', text_color = '#fff', font = ('Century Gothic bold', 24), fg_color = 'teal').place(x = 190, y = 10)  #fonte e tamanho do texto
        
        text = ctk.CTkLabel(self, text = 'Por favor, preencha todos os campos de formulário!', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff']).place(x = 50, y = 70) #fazendo isso para que as cores se alterem quando o usuário quiser.
        
        ficheiro = pathlib.Path('Clientes.xlsx') #criando a pasta e folha de trabalho no excel 
        
        if ficheiro.exists(): #se o ficheiro existe
            pass
        else: #se não existir um ficheiro, um será criado.
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1']= 'Nome Completo'
            folha['B1'] = 'Nº para Contato'
            folha['C1'] = 'Idade'
            folha['D1'] = 'Gênero'
            folha['E1'] = 'Endereço'
            folha['F1'] = 'Observações'   
            ficheiro.save('Clientes.xlsx')
        
        #Botões salvar e limpar (commands)
        def submit():  #não precisa do self pois ele está dentro da classe e função mãe
            #função de salvar dados

                #colunas do excel
            nome = nome_valor.get() #deve ser usado assim, para não resultar em um erro
            num = num_valor.get()
            idade = idade_valor.get()
            endereço = endereço_valor.get()
            genero = genero_combobox.get()
            obs = obs_textbox.get(0.0, END) #pegar desde a linha e coluna 0 até o fim do textbox!

            """Agora, iremos para a parte do ficheiro, onde iremos guardar os dados do usuário no excel"""
            
            ficheiro = openpyxl.load_workbook('Clientes.xlsx') #este é só para leitura.
            folha = ficheiro.active
            folha.cell(column = 1, row = folha.max_row+1, value = nome)
            #só precisa deixar +1 pq já tem +1 linha adicional, automaticamente, nas próximas colunas, ele já vai entender que as informações estão nas outras linhas!
            
            folha.cell(column = 2, row = folha.max_row, value = num) #num = contato
            folha.cell(column = 3, row = folha.max_row, value = idade)
            folha.cell(column = 4, row = folha.max_row, value = endereço)
            folha.cell(column = 5, row = folha.max_row, value = genero)
            folha.cell(column = 6, row = folha.max_row, value = obs)
            
            ficheiro.save(r'Clientes.xlsx') #salvarei dentro neste workbook(folha de trabalho) que abri no excel!
            messagebox.showinfo('Sistema', 'Dados Salvos com Sucesso!')
            #primeira coluna. Esse mais 1 ai é para quando vc ir adicionando mais dados, ir sempre criando mais linhas.
        def clear():
            #função de limpar dados
            #todos eles devem estar como Stringvar, senão vc não consegue limpar!(deve transformá-los em variáveis de string)
            nome = nome_valor.set('') 
            num = num_valor.set('')
            idade = idade_valor.set('')
            endereço = endereço_valor.set('')
            obs = obs_textbox.delete(0.0, END) #aqui não precisa ser STRINGVAR já que o textbox de obs é uma textbox. Devemos pegar da linha e coluna 0 até o final e deletar tudo.
            #o set está com strings pois ele irá setar os dados em strings/dados vazios! O combobox não será necessário(é o de escolha)
            
        
        #Text Variables/ variáveis de texto 
        nome_valor = StringVar()
        num_valor = StringVar()
        idade_valor = StringVar()
        endereço_valor = StringVar()
        #Entrys principais
        nome_entry = ctk.CTkEntry(self, width = 350, font = ('Century Gothic bold', 16), fg_color = 'transparent', textvariable = nome_valor)
        #fg_color é cor de fundo
        
        numero_entry = ctk.CTkEntry(self, width = 200, font = ('Century Gothic bold', 16), fg_color = 'transparent', textvariable = num_valor)
        
        idade_entry = ctk.CTkEntry(self, width = 150, font = ('Century Gothic bold', 16), fg_color = 'transparent', textvariable = idade_valor)
        
        endereço_entry = ctk.CTkEntry(self, width = 200, font = ('Century Gothic bold', 16), fg_color = 'transparent', textvariable = endereço_valor)
        
        #ComboBox para posicionarmos o gênero do usuário
        
        genero_combobox = ctk.CTkComboBox(self, values = ['Masculino', 'Feminino'], font=('Century Gothic bold', 14), border_color = '#aaa', border_width = 2, width = 150)
        genero_combobox.set('Masculino')
        
        
        
        #TextBox de observações
        
        obs_textbox = ctk.CTkTextbox(self, width = 470, height = 150, font = ('arial ', 18), border_color = '#aaa', border_width = 2)
        
            
        #labels principais
        
        lb_nome = ctk.CTkLabel(self, text = 'Nome Completo:', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff']) 
    
        lb_numero = ctk.CTkLabel(self, text = 'Nº para Contato:', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff']) 
        
        lb_idade = ctk.CTkLabel(self, text = 'Idade', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff']) 
        
        lb_genero = ctk.CTkLabel(self, text = 'Gênero', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff']) 
        
        lb_endereço = ctk.CTkLabel(self, text = 'Endereço', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff']) 
        
        lb_obs = ctk.CTkLabel(self, text = 'Observações:', font = ('Century Gothic bold', 16),
        text_color = ['#000', '#fff'])
    
    
        #Botões de salvar e limpar dadods
    
        btn_clique = ctk.CTkButton(self, text = 'Salvar Dados'.upper(), command = submit, fg_color = '#151', hover_color = '#131').place(x = 300, y = 420) #fazendo o botão de salvar dados
        btn_limpar = ctk.CTkButton(self, text = 'Limpar Dados'.upper(), command = clear, fg_color = '#555', hover_color = '#333').place(x = 500,y = 420)
        #mexendo na estética da cor normal e na cor quando passamos o mouse por cima.
        
        
        
        #Posicionamento de elementos na janela
        
        lb_nome.place(x = 50, y =120)
        nome_entry.place(x = 50, y = 150)
        
        lb_numero.place(x = 450, y = 120)
        numero_entry.place(x = 450, y = 150)
        
        lb_idade.place(x = 300, y = 190)
        idade_entry.place(x = 300, y = 220)
        
        lb_genero.place(x = 500, y = 190)
        genero_combobox.place(x = 500, y = 220)
        
        lb_endereço.place(x = 50, y = 190)
        endereço_entry.place(x = 50, y = 220)
        
        lb_obs.place(x = 50, y = 260)
        obs_textbox.place(x = 180, y = 270)
        
    
    
    
    def change_apm(self, nova_aparencia): #habilitando a troca de aparência
        ctk.set_appearance_mode(nova_aparencia)




if __name__ == '__main__': #fazendo a janela rodar
    app = App()
    app.mainloop()